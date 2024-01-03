import openpyxl

save_path = '03.excel_auto/test_data.xlsx'

# 기존 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_path)

# 활성화된 시트 선택
ws = wb.active

# 데이터 추가(1)
ws['A1'] = '날짜'
ws['B1'] = '제품명'
ws['C1'] = '가격'
ws['D1'] = '수량'
ws['E1'] = '합계'

# 데이터 추가(2)
ws.cell(row=2, column=1, value='2024-01-01') # A2
ws.cell(row=2, column=2, value='키보드') # B2
ws.cell(row=2, column=3, value=100000) # C2
ws.cell(row=2, column=4, value=10) # D2
ws.cell(row=2, column=5, value='=C2*D2') # E2

# 데이터 추가(3)
ws.append(['2023-01-02','기계식 키보드', 12000,15, '=C3*D3'])

# 데이터 수정
ws['C2'] = 40000
ws['D2'] = 50

# 데이터 삭제
del ws['A3']

# 엑셀 저장
wb.save(save_path)